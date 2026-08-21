"""
Converts a filled-in Positions.xlsx into data/jobs.json for the job board.

Usage:
    python scripts/import_jobs.py Positions.xlsx

Optional second argument to write somewhere else (defaults to data/jobs.json):
    python scripts/import_jobs.py Positions.xlsx path/to/output.json

Behavior:
- Reads the "Postings" sheet.
- Row 2 is skipped automatically if it's still the shipped example row
  (detected by its italic formatting, not its content — so real data that
  happens to match the example text is never dropped). Any other fully
  blank row is skipped too.
- Rows need at least Role, Venue and City — anything missing one of those
  is skipped and reported.
- Rows with a blank ID get a new one assigned. Rows with an ID keep it
  (this is how editing an existing posting vs. adding a new one is told
  apart — don't hand-edit IDs).
- Status defaults to "open" if blank; anything other than "open"/"filled"
  is treated as "open" and flagged.
- Apply Method must be instagram, email, or link. If it's blank but Apply
  Contact isn't, the method is guessed from the contact's shape (looks like
  an email vs. everything else) and flagged. Anything else unrecognized
  defaults to link and is flagged.
- Posted Date defaults to today (UTC) if blank, for new rows; existing
  rows keep their original date unless you type a new one.
- The output REPLACES the target JSON file entirely — the spreadsheet is
  treated as the full, current list of postings, not a diff.
"""

import sys
import json
import uuid
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

HEADERS = ["ID", "Role", "Venue", "City", "Country", "Description", "Apply Method", "Apply Contact", "Status", "Posted Date"]
VALID_STATUS = {"open", "filled"}
VALID_APPLY_METHOD = {"instagram", "email", "link"}


def normalize_date(value) -> str:
    if value is None:
        return date.today().isoformat()
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return date.today().isoformat()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_jobs.py <input.xlsx> [output.json]")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("data/jobs.json")

    wb = load_workbook(input_path, data_only=True)
    if "Postings" not in wb.sheetnames:
        print('Error: no "Postings" sheet found in this file.')
        sys.exit(1)
    ws = wb["Postings"]

    # map header row -> column index, in case columns were reordered
    header_row = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        name = clean(cell.value)
        if name:
            header_row[name] = col_idx

    missing_headers = [h for h in HEADERS if h not in header_row]
    if missing_headers:
        print(f"Error: missing expected column(s): {', '.join(missing_headers)}")
        sys.exit(1)

    def get(row, name):
        return row[header_row[name] - 1].value

    jobs = []
    skipped = []
    warnings = []

    for row in ws.iter_rows(min_row=2):
        excel_row_num = row[0].row

        role = clean(get(row, "Role"))
        venue = clean(get(row, "Venue"))
        city = clean(get(row, "City"))
        description = clean(get(row, "Description"))
        apply_method_raw = clean(get(row, "Apply Method")).lower()
        apply_contact = clean(get(row, "Apply Contact"))
        country = clean(get(row, "Country"))
        status_raw = clean(get(row, "Status")).lower()
        job_id = clean(get(row, "ID"))
        posted_raw = get(row, "Posted Date")
        instagram_url = clean(get(row, "Instagram URL")) if "Instagram URL" in header_row else ""

        # skip fully blank rows
        if not any([role, venue, city, description, apply_method_raw, apply_contact, country, status_raw, job_id]):
            continue

        # skip the still-unedited example row: require BOTH the italic
        # example styling AND the exact shipped example text, so real
        # postings are never dropped just for reusing similar wording or
        # inheriting the row's formatting after being typed over it
        role_cell = row[header_row["Role"] - 1]
        is_example_styled = bool(role_cell.font and role_cell.font.italic)
        is_example_text = role == "Line Cook" and venue == "Alba Restaurant"
        if is_example_styled and is_example_text and not job_id:
            continue

        missing = [name for name, val in [("Role", role), ("Venue", venue), ("City", city)] if not val]
        if missing:
            skipped.append(f"row {excel_row_num}: missing {', '.join(missing)}")
            continue

        status = status_raw if status_raw in VALID_STATUS else "open"
        if status_raw and status_raw not in VALID_STATUS:
            warnings.append(f"row {excel_row_num}: unrecognized status \"{status_raw}\", defaulted to open")

        if apply_method_raw in VALID_APPLY_METHOD:
            apply_method = apply_method_raw
        elif apply_method_raw:
            apply_method = "link"
            warnings.append(f"row {excel_row_num}: unrecognized apply method \"{apply_method_raw}\", defaulted to link")
        elif apply_contact:
            # no method given but a contact was — take a best guess so the
            # posting isn't silently missing its apply info
            if "@" in apply_contact and "." in apply_contact.split("@")[-1]:
                apply_method = "email"
            else:
                apply_method = "instagram"
            warnings.append(f"row {excel_row_num}: Apply Method left blank, guessed \"{apply_method}\" from Apply Contact")
        else:
            apply_method = "instagram"

        jobs.append({
            "id": job_id or str(uuid.uuid4()),
            "role": role,
            "venue": venue,
            "city": city,
            "country": country,
            "description": description,
            "applyMethod": apply_method,
            "applyContact": apply_contact,
            "instagramUrl": instagram_url,
            "status": status,
            "postedAt": normalize_date(posted_raw),
        })

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(jobs, indent=2) + "\n", encoding="utf-8")

    print(f"Wrote {len(jobs)} posting(s) to {output_path}")
    if skipped:
        print(f"\nSkipped {len(skipped)} row(s):")
        for s in skipped:
            print(f"  - {s}")
    if warnings:
        print(f"\n{len(warnings)} warning(s):")
        for w in warnings:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
