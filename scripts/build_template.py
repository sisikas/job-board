"""
Generates Positions.xlsx — the fillable spreadsheet template for job
postings. Run this once to (re)create the template:

    python scripts/build_template.py

Not needed for day-to-day use — only re-run it if you want to change the
template's columns or instructions.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_PATH = "Positions.xlsx"

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="171717")
BODY_FONT = Font(name="Arial", size=11)
EXAMPLE_FONT = Font(name="Arial", size=11, italic=True, color="9CA3AF")
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("ID", 10, "Leave blank for a new posting. Do not edit for existing postings — it's used to update the right row."),
    ("Role", 22, "Job title, e.g. Line Cook, Bartender, Front of House Manager."),
    ("Venue", 22, "Restaurant / company / venue name."),
    ("City", 16, "e.g. Athens"),
    ("Country", 16, "e.g. Greece"),
    ("Description", 45, "One or two sentences applicants will see."),
    ("Apply Method", 16, "instagram, email, or link — how applicants should get in touch."),
    ("Apply Contact", 26, "Matches Apply Method: an @handle for instagram, an address for email, or a URL for link."),
    ("Instagram URL", 28, "The business's Instagram — a full URL or @handle. Shown as an icon on the job page."),
    ("Status", 12, "open or filled. Filled postings are hidden from the public search."),
    ("Posted Date", 14, "Optional, format YYYY-MM-DD. Leave blank to use today's date for new postings."),
]

EXAMPLE_ROW = [
    "",
    "Line Cook",
    "Alba Restaurant",
    "New York",
    "United States",
    "Busy Italian kitchen looking for an experienced line cook for the dinner shift.",
    "instagram",
    "your_hiring_account",
    "https://instagram.com/alba",
    "open",
    "",
]


def build():
    wb = Workbook()

    # --- Instructions sheet ---
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_info.sheet_view.showGridLines = False
    ws_info.column_dimensions["A"].width = 90

    lines = [
        ("How to use this file", True),
        ("", False),
        ("1. Open the “Postings” tab.", False),
        ("2. Row 2 (in italics) is an example — delete it once you understand the format.", False),
        ("3. Add one row per job posting. Fill in Role, Venue and City at minimum.", False),
        ("4. Set Apply Method to instagram, email, or link, and put the matching contact in", False),
        ("   Apply Contact — an @handle, an email address, or a URL.", False),
        ("5. Leave the ID column blank for anything new — it gets assigned automatically.", False),
        ("6. To take a posting down, set its Status to “filled” rather than deleting the row —", False),
        ("   that keeps its ID stable if you reopen it later.", False),
        ("7. Save the file and send it back so it can be turned into the live listing.", False),
        ("", False),
        ("Columns", True),
        ("", False),
    ] + [(f"• {name}: {help_text}", False) for name, _, help_text in COLUMNS]

    r = 1
    for text, is_heading in lines:
        cell = ws_info.cell(row=r, column=1, value=text)
        cell.font = Font(name="Arial", size=13, bold=True) if is_heading else Font(name="Arial", size=11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # --- Postings sheet ---
    ws = wb.create_sheet("Postings")
    ws.sheet_view.showGridLines = False

    for col_idx, (name, width, _help) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 22

    # example row (row 2)
    for col_idx, value in enumerate(EXAMPLE_ROW, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = EXAMPLE_FONT
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 40

    # a handful of blank, pre-formatted rows ready to fill in
    for row_idx in range(3, 30):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # dropdown validations, columns located dynamically so they stay correct
    # if COLUMNS is reordered
    col_letter = {name: get_column_letter(i) for i, (name, _, _) in enumerate(COLUMNS, start=1)}

    dv_method = DataValidation(type="list", formula1='"instagram,email,link"', allow_blank=True)
    ws.add_data_validation(dv_method)
    dv_method.add(f"{col_letter['Apply Method']}2:{col_letter['Apply Method']}500")

    dv_status = DataValidation(type="list", formula1='"open,filled"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"{col_letter['Status']}2:{col_letter['Status']}500")

    ws.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    build()
